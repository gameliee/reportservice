"""read, valid excel files"""
from io import BytesIO
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from motor.motor_asyncio import AsyncIOMotorDatabase

from .models import AppConst, ExcelInvalidException
from ..stat import get_dataframe


def read_excel_validate(excel_bytes: bytes) -> pd.DataFrame | None:
    """Read the excel files stored as byte stream
    The excel file must follow requirements:
    - The heading is at the row 3
    - Must have column named "Mã nhân viên"
    - Columns with following name will automatically be filled:
        - "ghi nhận lần đầu"
        - "ghi nhận lần cuối"
        - "trạng thái mẫu"

    The function returns a Pandas Dataframe, with **all columns** in the same order as the excel file.
    By keeping all the columns the same, the dataframe could be use to write back to the excel file
    with minimum format changes.

    Args:
        excel_bytes (bytes): bytestream of the excel file. Can be provide as
        ```
        with open('example.xlsx', 'rb') as f:
            excel_bytes = f.read()
        ```

    Returns:
        pd.DataFrame | None: return None if cannot read the bytestream, or required columns are missing
    """
    try:
        file_like_excel = BytesIO(excel_bytes)
        df = pd.read_excel(file_like_excel, header=2, dtype=str, engine="openpyxl")
        df.columns = [x.lower().strip() for x in df.columns]
    except Exception as e:  # noqa: F841
        # TODO: log the error here
        print("could not read excel file stored")
        df = pd.DataFrame(columns=[AppConst.ESTAFF])

    # only care about defined columns
    if AppConst.ESTAFF not in df.columns:
        raise ExcelInvalidException(f'in excel file, column "{AppConst.ESTAFF}" not found')

    return df


def fill_excel(excel_bytes: bytes, filldata: pd.DataFrame) -> bytes | None:
    """fill the data from `filldata` into the excel file stored in `excel_bytes`
    `filldate` must have all the same columns, with same order, as `excel_bytes`
    By doing this, all the cell format in excel_bytes will not be changes.

    Args:
        excel_bytes (bytes): bytestream of the excel file. Can be provide as
        ```
        with open('example.xlsx', 'rb') as f:
            excel_bytes = f.read()
        ```
        filldata (pd.DataFrame): data to fill into excel_bytes

    Returns:
        bytes | None: return None if there is something wrong
    """
    # Validate the `excel_bytes` and `filldata` has the same columns
    df = read_excel_validate(excel_bytes)
    if not df.columns.equals(filldata.columns):
        return None

    file_like_object = BytesIO(excel_bytes)
    wb = load_workbook(file_like_object)
    ws = wb.active  # gets first sheet

    rows = dataframe_to_rows(filldata, index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            # TODO: only modify columns in interested columns
            ws.cell(row=r_idx + 3, column=c_idx, value=value)

    # clean remainding rows
    for row in ws.iter_rows(min_row=len(filldata) + 4):
        for cell in row:
            cell.value = None
            cell.style = "Normal"

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()


def excel_to_html(excelbytes: bytes) -> str:
    """given the excel file, convert it to html using pandas"""
    df = read_excel_validate(excel_bytes=excelbytes)
    return df.fillna("").to_html(index=False)


async def extract_and_fill_excel(
    db: AsyncIOMotorDatabase,
    staff_collection: str,
    bodyfacename_collection: str,
    excelbytes: bytes,
    begin: datetime,
    end: datetime,
) -> bytes:
    """given an excel file, fill it with data from database. Only following columns will be filled:
    - staffcode
    - first recognition time
    - last recognition time
    - has_sample

    Args:
        db (AsyncIOMotorDatabase): database
        staff_collection (str): collection name of staffs
        bodyfacename_collection (str): collection name of BodyFaceName
        excelbytes (bytes): bytestream of the excel file. Can be provide as
        ```
        with open('example.xlsx', 'rb') as f:
            excel_bytes = f.read()
        ```
        begin (datetime): begin of the query window
        end (datetime): end of the query window

    Returns:
        bytes: result excel file as bytestream
    """
    origin_df = read_excel_validate(excelbytes)
    staffcodes = origin_df[AppConst.ESTAFF].dropna().to_list()

    result_df = await get_dataframe(db, staff_collection, bodyfacename_collection, staffcodes, begin, end)

    # Update the null elements from result_df into origin_df
    result_df = result_df.merge(
        origin_df[AppConst.ESTAFF], left_on=AppConst.ESTAFF, right_on=AppConst.ESTAFF, how="right"
    )  # keep result_df and origin_df same index
    origin_columns = origin_df.columns  # keep columns order
    df = origin_df.combine_first(result_df)  # where magics happen
    df = df.reindex(columns=origin_columns)  # restore columns order

    outbytes = fill_excel(excelbytes, df)
    return outbytes
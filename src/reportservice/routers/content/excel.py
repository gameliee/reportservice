"""read, valid excel files"""
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fastapi.encoders import jsonable_encoder
from ..stat import PersonInoutCollection
from .models import ExcelColumn, ExcelInvalidException


def read_excel_validate(excel_bytes: bytes) -> pd.DataFrame | None:
    """Read the excel files stored as byte stream
    The excel file must follow requirements:
    - The heading is at the row 3
    - Must have column named "Mã nhân viên"
    - Columns with following name will automatically be filled:
        - "ghi nhận lần đầu"
        - "ghi nhận lần cuối"
        - "trạng thái mẫu"

    The function returns a Pandas Dataframe, with **all columns** in the same order as the excel file.
    By keeping all the columns the same, the dataframe could be use to write back to the excel file
    with minimum format changes.

    Args:
        excel_bytes (bytes): bytestream of the excel file. Can be provide as
        ```
        with open('example.xlsx', 'rb') as f:
            excel_bytes = f.read()
        ```

    Returns:
        pd.DataFrame | None: return None if cannot read the bytestream, or required columns are missing
    """
    try:
        file_like_excel = BytesIO(excel_bytes)
        df = pd.read_excel(file_like_excel, header=2, dtype=str, engine="openpyxl")
        df.columns = [x.lower().strip() for x in df.columns]
    except Exception as e:  # noqa: F841
        # TODO: log the error here
        print("could not read excel file stored")
        df = pd.DataFrame(columns=[ExcelColumn.ESTAFF])

    # only care about defined columns
    if ExcelColumn.ESTAFF not in df.columns:
        raise ExcelInvalidException(f'in excel file, column "{ExcelColumn.ESTAFF}" not found')

    return df


def fill_excel(excel_bytes: bytes, filldata: pd.DataFrame) -> bytes | None:
    """fill the data from `filldata` into the excel file stored in `excel_bytes`
    `filldate` must have all the same columns, with same order, as `excel_bytes`
    By doing this, all the cell format in excel_bytes will not be changes.

    Args:
        excel_bytes (bytes): bytestream of the excel file. Can be provide as
        ```
        with open('example.xlsx', 'rb') as f:
            excel_bytes = f.read()
        ```
        filldata (pd.DataFrame): data to fill into excel_bytes

    Returns:
        bytes | None: return None if there is something wrong
    """
    # Validate the `excel_bytes` and `filldata` has the same columns
    df = read_excel_validate(excel_bytes)
    if not df.columns.equals(filldata.columns):
        return None

    file_like_object = BytesIO(excel_bytes)
    wb = load_workbook(file_like_object)
    ws = wb.active  # gets first sheet

    rows = dataframe_to_rows(filldata, index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            # TODO: only modify columns in interested columns
            ws.cell(row=r_idx + 3, column=c_idx, value=value)

    # clean remainding rows
    for row in ws.iter_rows(min_row=len(filldata) + 4):
        for cell in row:
            cell.value = None
            cell.style = "Normal"

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()


def excel_to_html(excelbytes: bytes) -> str:
    """given the excel file, convert it to html using pandas"""
    df = read_excel_validate(excel_bytes=excelbytes)
    return df.fillna("").to_html(index=False)


def fill_personinout_to_excel(
    people_inout: PersonInoutCollection,
    excelbytes: bytes,
) -> bytes:
    """Fill excel file with PersonInout
    In the excel file, only following columns will be filled:
    - staffcode
    - first recognition time
    - last recognition time
    - has_sample

    Args:
        people_inout (PersonInoutCollection): the data to fill into excel file
        excelbytes (bytes): bytestream of the excel file. Can be provide as
        ```
        with open('example.xlsx', 'rb') as f:
            excel_bytes = f.read()
        ```

    Returns:
        bytes: result excel file as bytestream
    """
    origin_df = read_excel_validate(excelbytes)

    if people_inout.count == 0:
        return excelbytes

    result_df = pd.DataFrame(jsonable_encoder(people_inout.values))
    # TODO: ensure column name
    result_df.rename(
        columns={
            "staff_code": ExcelColumn.ESTAFF,
            "first_record": ExcelColumn.EFIRST,
            "last_record": ExcelColumn.ELASTT,
            "sample_state": ExcelColumn.ESAMPL,
        },
        inplace=True,
    )

    # Update the null elements from result_df into origin_df
    result_df = result_df.merge(
        origin_df[ExcelColumn.ESTAFF], left_on=ExcelColumn.ESTAFF, right_on=ExcelColumn.ESTAFF, how="right"
    )  # keep result_df and origin_df same index
    origin_columns = origin_df.columns  # keep columns order
    df = origin_df.combine_first(result_df)  # where magics happen
    df = df.reindex(columns=origin_columns)  # restore columns order

    outbytes = fill_excel(excelbytes, df)
    return outbytes


def convert_personinout_to_excel(
    people_inout: PersonInoutCollection,
) -> bytes:
    if people_inout.count == 0:
        # create empty df with columns from ExcelColumn
        empty_df = pd.DataFrame(
            columns=[ExcelColumn.ESTAFF, ExcelColumn.EFIRST, ExcelColumn.ELASTT, ExcelColumn.ESAMPL]
        )
        # convert to excel
        virtual_workbook = BytesIO()
        empty_df.to_excel(virtual_workbook, index=False)
        return virtual_workbook.getvalue()

    result_df = pd.DataFrame(jsonable_encoder(people_inout.values))
    result_df.fillna("", inplace=True)
    result_df.rename(
        columns={
            "staff_code": ExcelColumn.ESTAFF,
            "first_record": ExcelColumn.EFIRST,
            "last_record": ExcelColumn.ELASTT,
            "sample_state": ExcelColumn.ESAMPL,
        },
        inplace=True,
    )

    # Sort by custom rules
    sort_by = []
    if "unit" in result_df.columns:
        result_df["unit_sort"] = result_df["unit"].str.lower().apply(lambda x: "0" + x if "giám đốc" in x else x)
        sort_by.append("unit_sort")

    if "department" in result_df.columns:
        result_df["department_sort"] = result_df["department"].str.lower()
        sort_by.append("department_sort")

    if "title" in result_df.columns:

        def _sort_title(x):
            if x.startswith("chánh"):
                return "0" + x
            elif x.startswith("trưởng"):
                return "1" + x
            elif x.startswith("phó"):
                return "2" + x
            else:
                return x

        result_df["title_sort"] = result_df["title"].str.lower().apply(_sort_title)
        sort_by.append("title_sort")

    result_df.sort_values(by=sort_by + [ExcelColumn.ESTAFF], inplace=True)
    # drop sort columns
    result_df.drop(columns=sort_by, inplace=True)

    # create index column
    result_df.index.name = "STT"
    result_df.index.name = "STT"
    result_df.reset_index(inplace=True)

    # reorder the columns. First columns should be STT, staffcode, first, last, sample. And then the rest
    columns = result_df.columns.tolist()
    columns.remove("STT")
    columns.remove(ExcelColumn.ESTAFF)
    columns.remove(ExcelColumn.EFIRST)
    columns.remove(ExcelColumn.ELASTT)
    columns.remove(ExcelColumn.ESAMPL)
    columns = ["STT", ExcelColumn.ESTAFF, ExcelColumn.EFIRST, ExcelColumn.ELASTT, ExcelColumn.ESAMPL] + columns
    result_df = result_df[columns]

    # convert to excel
    virtual_workbook = BytesIO()
    result_df.to_excel(virtual_workbook, startrow=2, index=False)  # startrow=2 is fixed to match input excel file
    return virtual_workbook.getvalue()
